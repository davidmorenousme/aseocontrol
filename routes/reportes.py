from flask import Blueprint, session, redirect, url_for, send_file
from datetime import date, datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

reportes = Blueprint('reportes', __name__)
mysql = None

def init_mysql(mysql_instance):
    global mysql
    mysql = mysql_instance

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated

# Estilos reutilizables
def estilo_encabezado():
    return Font(bold=True, color='FFFFFF', size=11)

def fill_verde():
    return PatternFill('solid', fgColor='2d7a4f')

def fill_verde_claro():
    return PatternFill('solid', fgColor='e8f5ee')

def fill_gris():
    return PatternFill('solid', fgColor='f5f7f4')

def borde():
    lado = Side(style='thin', color='dce8e0')
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_encabezados(ws, columnas):
    for col, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = estilo_encabezado()
        cell.fill = fill_verde()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde()
    ws.row_dimensions[1].height = 25


# ── 1. Lista completa de empleados ──
@reportes.route('/admin/exportar/empleados')
@login_required
def exportar_empleados():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.documento, e.cargo, e.telefono,
               e.tipo, c.nombre as conjunto,
               CASE WHEN e.activo = 1 THEN 'Activo' ELSE 'Inactivo' END as estado,
               h.hora_entrada, h.hora_salida
        FROM empleados e
        LEFT JOIN conjuntos c ON c.id = e.conjunto_id
        LEFT JOIN horarios h ON h.empleado_id = e.id
        ORDER BY e.activo DESC, c.nombre, e.nombre
    """)
    empleados = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Empleados'

    columnas = ['Nombre', 'Documento', 'Cargo', 'Teléfono', 'Tipo', 'Conjunto', 'Estado', 'Hora Entrada', 'Hora Salida']
    aplicar_encabezados(ws, columnas)

    for i, e in enumerate(empleados, 2):
        fila = [
            e['nombre'], e['documento'], e['cargo'], e['telefono'],
            e['tipo'].capitalize() if e['tipo'] else '',
            e['conjunto'] or 'Sin conjunto',
            e['estado'],
            str(e['hora_entrada'])[:5] if e['hora_entrada'] else '—',
            str(e['hora_salida'])[:5] if e['hora_salida'] else '—',
        ]
        fill = fill_verde_claro() if i % 2 == 0 else fill_gris()
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill = fill
            cell.border = borde()
            cell.alignment = Alignment(vertical='center')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f'empleados_{date.today()}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 2. Empleados con novedades activas ──
@reportes.route('/admin/exportar/novedades')
@login_required
def exportar_novedades():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.documento, c.nombre as conjunto,
               n.tipo, n.fecha_inicio, n.fecha_fin, n.descripcion
        FROM novedades n
        JOIN empleados e ON e.id = n.empleado_id
        LEFT JOIN conjuntos c ON c.id = e.conjunto_id
        WHERE n.fecha_fin >= %s OR n.fecha_fin IS NULL
        ORDER BY n.tipo, e.nombre
    """, (date.today(),))
    novedades = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Novedades Activas'

    columnas = ['Empleado', 'Documento', 'Conjunto', 'Tipo', 'Fecha Inicio', 'Fecha Fin', 'Descripción']
    aplicar_encabezados(ws, columnas)

    for i, n in enumerate(novedades, 2):
        fila = [
            n['nombre'], n['documento'], n['conjunto'] or '—',
            n['tipo'].replace('_', ' ').capitalize(),
            n['fecha_inicio'].strftime('%d/%m/%Y') if n['fecha_inicio'] else '—',
            n['fecha_fin'].strftime('%d/%m/%Y') if n['fecha_fin'] else 'Indefinido',
            n['descripcion'] or '—',
        ]
        fill = fill_verde_claro() if i % 2 == 0 else fill_gris()
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill = fill
            cell.border = borde()
            cell.alignment = Alignment(vertical='center')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f'novedades_{date.today()}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 3. Historial de retiros ──
@reportes.route('/admin/exportar/retiros')
@login_required
def exportar_retiros():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.documento, c.nombre as conjunto,
               n.fecha_inicio as fecha_retiro, n.descripcion
        FROM novedades n
        JOIN empleados e ON e.id = n.empleado_id
        LEFT JOIN conjuntos c ON c.id = e.conjunto_id
        WHERE n.tipo = 'retiro'
        ORDER BY n.fecha_inicio DESC
    """)
    retiros = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Retiros'

    columnas = ['Empleado', 'Documento', 'Conjunto', 'Fecha Retiro', 'Motivo']
    aplicar_encabezados(ws, columnas)

    for i, r in enumerate(retiros, 2):
        fila = [
            r['nombre'], r['documento'], r['conjunto'] or '—',
            r['fecha_retiro'].strftime('%d/%m/%Y') if r['fecha_retiro'] else '—',
            r['descripcion'] or '—',
        ]
        fill = fill_verde_claro() if i % 2 == 0 else fill_gris()
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill = fill
            cell.border = borde()
            cell.alignment = Alignment(vertical='center')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f'retiros_{date.today()}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 4. Reporte mensual de asistencia ──
@reportes.route('/admin/exportar/mensual/<int:anio>/<int:mes>')
@login_required
def exportar_mensual(anio, mes):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.documento, c.nombre as conjunto,
               DATE(r.timestamp) as fecha,
               MAX(CASE WHEN r.tipo = 'entrada'          THEN TIME(r.timestamp) END) as entrada,
               MAX(CASE WHEN r.tipo = 'salida_break'     THEN TIME(r.timestamp) END) as sal_break,
               MAX(CASE WHEN r.tipo = 'regreso_break'    THEN TIME(r.timestamp) END) as reg_break,
               MAX(CASE WHEN r.tipo = 'salida_almuerzo'  THEN TIME(r.timestamp) END) as sal_almuerzo,
               MAX(CASE WHEN r.tipo = 'regreso_almuerzo' THEN TIME(r.timestamp) END) as reg_almuerzo,
               MAX(CASE WHEN r.tipo = 'salida'           THEN TIME(r.timestamp) END) as salida
        FROM registros r
        JOIN empleados e ON e.id = r.empleado_id
        LEFT JOIN conjuntos c ON c.id = e.conjunto_id
        WHERE YEAR(r.timestamp) = %s AND MONTH(r.timestamp) = %s
        GROUP BY e.id, e.nombre, e.documento, c.nombre, DATE(r.timestamp)
        ORDER BY c.nombre, e.nombre, fecha
    """, (anio, mes))
    registros = cur.fetchall()
    cur.close()

    meses = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{meses[mes]} {anio}'

    columnas = ['Empleado', 'Documento', 'Conjunto', 'Fecha',
                'Entrada', 'Sal. Break', 'Reg. Break',
                'Sal. Almuerzo', 'Reg. Almuerzo', 'Salida']
    aplicar_encabezados(ws, columnas)

    for i, r in enumerate(registros, 2):
        fila = [
            r['nombre'], r['documento'], r['conjunto'] or '—',
            r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '—',
            str(r['entrada'])[:5] if r['entrada'] else '—',
            str(r['sal_break'])[:5] if r['sal_break'] else '—',
            str(r['reg_break'])[:5] if r['reg_break'] else '—',
            str(r['sal_almuerzo'])[:5] if r['sal_almuerzo'] else '—',
            str(r['reg_almuerzo'])[:5] if r['reg_almuerzo'] else '—',
            str(r['salida'])[:5] if r['salida'] else '—',
        ]
        fill = fill_verde_claro() if i % 2 == 0 else fill_gris()
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.fill = fill
            cell.border = borde()
            cell.alignment = Alignment(vertical='center')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'asistencia_{meses[mes].lower()}_{anio}.xlsx'
    return send_file(output, download_name=nombre_archivo,
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')